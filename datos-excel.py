import tkinter as tk
from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

root=Tk()
root.title("Ingreso de Datos")
root.geometry('700x400+300+200')
root.resizable(False,False)
root.configure(bg="#326273")

archivo = pathlib.Path('datos.xslx')
if archivo.exists():
    pass
else:
    archivo=Workbook()
    hoja=archivo.active
    hoja['A1']='Nombre'
    hoja['B1']='Teléfono'
    hoja['C1']='Edad'
    hoja['D1']='Género'
    hoja['E1']='Dirección'

    archivo.save('datos.xlsx')


def enviar():
    nombre = nombreValor.get()
    numero = numeroValor.get()
    edad = edadValor.get()
    genero = genero_combobox.get()
    direccion = direccionEntrada.get(1.0, END)

    archivo = openpyxl.load_workbook('datos.xlsx')
    hoja = archivo.active
    hoja.cell(column=1, row=hoja.max_row+1, value=nombre)
    hoja.cell(column=2, row=hoja.max_row, value=numero)
    hoja.cell(column=3, row=hoja.max_row, value=edad)
    hoja.cell(column=4, row=hoja.max_row, value=genero)
    hoja.cell(column=5, row=hoja.max_row, value=direccion)

    archivo.save(r'datos.xlsx')

    messagebox.showinfo('Info', 'Datos guardados!')
    
    limpiar()


def limpiar():
    nombreValor.set('')
    numeroValor.set('')
    edadValor.set('')
    direccionEntrada.delete(1.0, END)
    nombreEntrada.focus_set()

#icon
#icon_image=PhotoImage(file="logo.png")
#root.iconphoto(False, icon_image)

#heading
Label(root, text="Por favor ingrese los datos.", font="arial 13", bg="#326273", fg="#fff").place(x=20, y=20)

#label
Label(root, text="Nombre", font=23, bg="#326273", fg="#fff").place(x=50, y=100)
Label(root, text="Número", font=23, bg="#326273", fg="#fff").place(x=50, y=150)
Label(root, text="Edad", font=23, bg="#326273", fg="#fff").place(x=50, y=200)
Label(root, text="Género", font=23, bg="#326273", fg="#fff").place(x=370, y=200)
Label(root, text="Dirección", font=23, bg="#326273", fg="#fff").place(x=50, y=250)

#Entradas
nombreValor = StringVar()
numeroValor = StringVar()
edadValor = StringVar()

nombreEntrada = Entry(root, textvariable=nombreValor, width=45, bd=2, font=20)
numeroEntrada = Entry(root, textvariable=numeroValor, width=45, bd=2, font=20)
edadEntrada = Entry(root, textvariable=edadValor, width=15, bd=2, font=20)


#genero
genero_combobox = Combobox(root, values=['Hombre', 'Mujer'], font='arial 14', state='r', width=14)
genero_combobox.set('Hombre')

direccionEntrada = Text(root, width=50, height=4, bd=2)

nombreEntrada.place(x=200, y=100)
numeroEntrada.place(x=200, y=150)
edadEntrada.place(x=200, y=200)
genero_combobox.place(x=440, y=200)
direccionEntrada.place(x=200, y=250)

nombreEntrada.focus_set()

#botones
Button(root, text='Guardar', bg='#326273', fg='white', width=15, height=2, command=enviar).place(x=200, y=350)
Button(root, text='Limpiar', bg='#326273', fg='white', width=15, height=2, command=limpiar).place(x=340, y=350)
Button(root, text='Salir', bg='#326273', fg='white', width=15, height=2, command=lambda:root.destroy()).place(x=480, y=350)

root.mainloop()
